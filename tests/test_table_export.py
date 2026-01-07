"""
Tests for Table Export Functionality

Tests CSV/Excel export, formatting, large datasets, and edge cases.
"""

import tempfile
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from kie.tables.builder import TableBuilder
from kie.tables.export import TableExporter, export_table
from kie.tables.schema import (
    Alignment,
    ColumnConfig,
    ColumnType,
    CurrencyFormat,
    PercentageFormat,
    TableConfig,
)


# ===== Fixtures =====


@pytest.fixture
def sample_data():
    """Sample table data."""
    return pd.DataFrame(
        {
            "region": ["North", "South", "East", "West"],
            "revenue": [1250000, 980000, 1450000, 1100000],
            "growth": [0.15, 0.08, 0.22, 0.12],
            "sales_count": [450, 320, 550, 410],
        }
    )


@pytest.fixture
def table_config(sample_data):
    """Basic table config."""
    builder = TableBuilder()
    return builder.build(sample_data, title="Sales Report")


@pytest.fixture
def table_config_with_totals(sample_data):
    """Table config with totals row."""
    builder = TableBuilder()
    config = builder.build(sample_data, title="Sales Report")
    builder.add_totals_row(config)
    return config


@pytest.fixture
def temp_dir():
    """Temporary directory for test files."""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


# ===== CSV Export Tests =====


def test_csv_export_basic(table_config, temp_dir):
    """Test basic CSV export."""
    exporter = TableExporter()
    output_path = temp_dir / "test.csv"

    result = exporter.to_csv(table_config, output_path, include_totals=False)

    assert result.exists()
    assert result.suffix == ".csv"

    # Read back and verify
    df = pd.read_csv(result)
    assert len(df) == 4
    assert list(df.columns) == ["Region", "Revenue", "Growth", "Sales Count"]


def test_csv_export_with_totals(table_config_with_totals, temp_dir):
    """Test CSV export includes totals row."""
    exporter = TableExporter()
    output_path = temp_dir / "test_totals.csv"

    result = exporter.to_csv(table_config_with_totals, output_path, include_totals=True)

    # Read back and verify
    df = pd.read_csv(result)
    assert len(df) == 5  # 4 data rows + 1 totals row

    # Check totals row
    totals_row = df.iloc[-1]
    assert totals_row["Region"] == "Total"
    assert totals_row["Revenue"] == 4780000  # Sum of revenue


def test_csv_export_hidden_columns(sample_data, temp_dir):
    """Test CSV export respects hidden columns."""
    builder = TableBuilder()
    config = builder.build(sample_data, title="Test")

    # Hide 'growth' column
    for col in config.columns:
        if col.key == "growth":
            col.hidden = True

    exporter = TableExporter()
    output_path = temp_dir / "test_hidden.csv"

    result = exporter.to_csv(config, output_path)

    df = pd.read_csv(result)
    assert "Growth" not in df.columns
    assert len(df.columns) == 3  # Only 3 visible columns


def test_csv_export_empty_data(temp_dir):
    """Test CSV export with empty data."""
    # Create DataFrame with columns but no rows
    data = pd.DataFrame(columns=["col1", "col2"])

    config = TableConfig(
        title="Empty",
        columns=[
            ColumnConfig(key="col1", header="Column 1"),
            ColumnConfig(key="col2", header="Column 2"),
        ],
        data=data.to_dict("records"),
    )

    exporter = TableExporter()
    output_path = temp_dir / "empty.csv"

    result = exporter.to_csv(config, output_path)

    df = pd.read_csv(result)
    assert len(df) == 0
    assert list(df.columns) == ["Column 1", "Column 2"]


def test_csv_export_creates_directories(table_config, temp_dir):
    """Test CSV export creates nested directories."""
    output_path = temp_dir / "subdir1" / "subdir2" / "test.csv"

    exporter = TableExporter()
    result = exporter.to_csv(table_config, output_path)

    assert result.exists()
    assert result.parent.parent.name == "subdir1"


# ===== Excel Export Tests =====


def test_excel_export_basic(table_config, temp_dir):
    """Test basic Excel export."""
    exporter = TableExporter()
    output_path = temp_dir / "test.xlsx"

    result = exporter.to_excel(table_config, output_path, style_headers=False)

    assert result.exists()
    assert result.suffix == ".xlsx"

    # Read back and verify
    df = pd.read_excel(result)
    assert len(df) == 4
    assert list(df.columns) == ["Region", "Revenue", "Growth", "Sales Count"]


def test_excel_export_with_styling(table_config, temp_dir):
    """Test Excel export with KDS header styling."""
    exporter = TableExporter()
    output_path = temp_dir / "styled.xlsx"

    result = exporter.to_excel(table_config, output_path, style_headers=True)

    # Load workbook and check styling
    wb = load_workbook(result)
    ws = wb.active

    # Check header row has KDS purple fill (openpyxl uses 00RRGGBB format)
    first_cell = ws["A1"]
    assert first_cell.fill.start_color.rgb == "007823DC"  # KDS purple

    # Check header font
    assert first_cell.font.bold is True
    assert first_cell.font.color.rgb == "00FFFFFF"  # White text


def test_excel_export_with_totals(table_config_with_totals, temp_dir):
    """Test Excel export includes totals row with bold font."""
    exporter = TableExporter()
    output_path = temp_dir / "totals.xlsx"

    result = exporter.to_excel(table_config_with_totals, output_path, style_headers=True)

    wb = load_workbook(result)
    ws = wb.active

    # Check totals row is bold
    totals_cell = ws[f"A{ws.max_row}"]
    assert totals_cell.font.bold is True
    assert totals_cell.value == "Total"


def test_excel_export_custom_sheet_name(table_config, temp_dir):
    """Test Excel export with custom sheet name."""
    exporter = TableExporter()
    output_path = temp_dir / "custom_sheet.xlsx"

    result = exporter.to_excel(table_config, output_path, sheet_name="Q4 Sales")

    wb = load_workbook(result)
    assert "Q4 Sales" in wb.sheetnames


def test_excel_export_column_width_adjustment(table_config, temp_dir):
    """Test Excel export auto-adjusts column widths."""
    exporter = TableExporter()
    output_path = temp_dir / "widths.xlsx"

    result = exporter.to_excel(table_config, output_path, style_headers=True)

    wb = load_workbook(result)
    ws = wb.active

    # Check column widths are adjusted (openpyxl default is ~8.43)
    # At minimum, width should be set (not None)
    col_width = ws.column_dimensions["A"].width
    assert col_width >= 8  # Should have a reasonable width


def test_excel_export_hidden_columns(sample_data, temp_dir):
    """Test Excel export respects hidden columns."""
    builder = TableBuilder()
    config = builder.build(sample_data, title="Test")

    # Hide 'growth' column
    for col in config.columns:
        if col.key == "growth":
            col.hidden = True

    exporter = TableExporter()
    output_path = temp_dir / "hidden.xlsx"

    result = exporter.to_excel(config, output_path)

    df = pd.read_excel(result)
    assert "Growth" not in df.columns


# ===== Totals Calculation Tests =====


def test_totals_calculation_sum():
    """Test totals row with sum aggregation."""
    data = pd.DataFrame({"value": [100, 200, 300]})

    config = TableConfig(
        title="Test",
        columns=[
            ColumnConfig(
                key="value", header="Value", type=ColumnType.NUMBER, footer_aggregate="sum"
            )
        ],
        data=data.to_dict("records"),
        show_totals_row=True,
    )

    exporter = TableExporter()
    totals = exporter._calculate_totals(config)

    assert totals["Value"] == 600


def test_totals_calculation_avg():
    """Test totals row with average aggregation."""
    data = pd.DataFrame({"value": [100, 200, 300]})

    config = TableConfig(
        title="Test",
        columns=[
            ColumnConfig(
                key="value", header="Value", type=ColumnType.NUMBER, footer_aggregate="avg"
            )
        ],
        data=data.to_dict("records"),
        show_totals_row=True,
    )

    exporter = TableExporter()
    totals = exporter._calculate_totals(config)

    assert totals["Value"] == 200


def test_totals_calculation_min_max():
    """Test totals row with min/max aggregation."""
    data = pd.DataFrame({"value": [100, 200, 50, 300]})

    config_min = TableConfig(
        title="Test",
        columns=[
            ColumnConfig(
                key="value", header="Value", type=ColumnType.NUMBER, footer_aggregate="min"
            )
        ],
        data=data.to_dict("records"),
        show_totals_row=True,
    )

    config_max = TableConfig(
        title="Test",
        columns=[
            ColumnConfig(
                key="value", header="Value", type=ColumnType.NUMBER, footer_aggregate="max"
            )
        ],
        data=data.to_dict("records"),
        show_totals_row=True,
    )

    exporter = TableExporter()
    totals_min = exporter._calculate_totals(config_min)
    totals_max = exporter._calculate_totals(config_max)

    assert totals_min["Value"] == 50
    assert totals_max["Value"] == 300


def test_totals_calculation_count():
    """Test totals row with count aggregation."""
    data = pd.DataFrame({"value": [100, None, 200, None, 300]})

    config = TableConfig(
        title="Test",
        columns=[
            ColumnConfig(
                key="value", header="Value", type=ColumnType.NUMBER, footer_aggregate="count"
            )
        ],
        data=data.to_dict("records"),
        show_totals_row=True,
    )

    exporter = TableExporter()
    totals = exporter._calculate_totals(config)

    assert totals["Value"] == 3  # Only counts non-null values


def test_totals_calculation_mixed_columns():
    """Test totals row with mixed aggregation types."""
    data = pd.DataFrame({"category": ["A", "B", "C"], "revenue": [100, 200, 300], "count": [5, 10, 15]})

    config = TableConfig(
        title="Test",
        columns=[
            ColumnConfig(key="category", header="Category", type=ColumnType.TEXT),
            ColumnConfig(
                key="revenue",
                header="Revenue",
                type=ColumnType.CURRENCY,
                footer_aggregate="sum",
            ),
            ColumnConfig(
                key="count", header="Count", type=ColumnType.NUMBER, footer_aggregate="sum"
            ),
        ],
        data=data.to_dict("records"),
        show_totals_row=True,
        totals_label="Grand Total",
    )

    exporter = TableExporter()
    totals = exporter._calculate_totals(config)

    assert totals["Category"] == "Grand Total"
    assert totals["Revenue"] == 600
    assert totals["Count"] == 30


# ===== Multi-Format Export Tests =====


def test_export_multiple_formats(table_config, temp_dir):
    """Test exporting to multiple formats at once."""
    result = export_table(table_config, temp_dir, formats=["csv", "excel"], base_name="sales")

    assert "csv" in result
    assert "excel" in result
    assert result["csv"].exists()
    assert result["excel"].exists()
    assert result["csv"].name == "sales.csv"
    assert result["excel"].name == "sales.xlsx"


def test_export_default_formats(table_config, temp_dir):
    """Test export with default formats (CSV and Excel)."""
    result = export_table(table_config, temp_dir)

    assert "csv" in result
    assert "excel" in result


def test_export_auto_base_name(table_config, temp_dir):
    """Test export generates base name from title."""
    result = export_table(table_config, temp_dir, formats=["csv"])

    assert result["csv"].name == "sales_report.csv"


def test_export_no_title_fallback(sample_data, temp_dir):
    """Test export uses 'table' as fallback when no title."""
    builder = TableBuilder()
    config = builder.build(sample_data)  # No title

    result = export_table(config, temp_dir, formats=["csv"])

    assert result["csv"].name == "table.csv"


def test_export_pdf_now_implemented(table_config, temp_dir):
    """Test PDF export is now implemented."""
    pytest = __import__('pytest')

    try:
        import reportlab  # noqa: F401
        # reportlab is installed - PDF export should work
        result = export_table(table_config, temp_dir, formats=["pdf"])
        assert "pdf" in result  # PDF is in output paths
        assert result["pdf"].exists()
        assert result["pdf"].suffix == ".pdf"
    except ImportError:
        # reportlab not installed - should skip gracefully
        result = export_table(table_config, temp_dir, formats=["pdf"])
        assert "pdf" not in result  # PDF not in output paths when reportlab missing


# ===== Large Dataset Tests =====


def test_export_large_dataset(temp_dir):
    """Test exporting a large dataset (10,000 rows)."""
    # Generate large dataset
    num_rows = 10000
    large_data = pd.DataFrame(
        {
            "id": range(num_rows),
            "category": (["Category A", "Category B", "Category C"] * (num_rows // 3 + 1))[:num_rows],
            "value": [i * 100 for i in range(num_rows)],
            "rate": [i / num_rows for i in range(num_rows)],
        }
    )

    builder = TableBuilder()
    config = builder.build(large_data, title="Large Dataset")

    # Export to CSV
    result = export_table(config, temp_dir, formats=["csv"])

    # Verify
    df = pd.read_csv(result["csv"])
    assert len(df) == 10000
    assert list(df.columns) == ["Id", "Category", "Value", "Rate"]


def test_export_large_dataset_excel(temp_dir):
    """Test exporting large dataset to Excel (performance check)."""
    # Generate 5,000 rows (Excel can handle this well)
    large_data = pd.DataFrame(
        {
            "id": range(5000),
            "value": [i * 100 for i in range(5000)],
        }
    )

    builder = TableBuilder()
    config = builder.build(large_data, title="Large Excel")

    # Export to Excel with styling
    exporter = TableExporter()
    output_path = temp_dir / "large.xlsx"

    result = exporter.to_excel(config, output_path, style_headers=True)

    # Verify
    df = pd.read_excel(result)
    assert len(df) == 5000


# ===== Edge Case Tests =====


def test_export_with_special_characters(temp_dir):
    """Test export handles special characters in data."""
    data = pd.DataFrame(
        {
            "name": ["O'Reilly", 'Smith & Co.', 'John "The Boss" Doe'],
            "notes": ["Test, test", "Line\nbreak", "Quote: 'test'"],
        }
    )

    builder = TableBuilder()
    config = builder.build(data, title="Special Chars")

    result = export_table(config, temp_dir, formats=["csv"])

    # Read back - pandas should handle special chars correctly
    df = pd.read_csv(result["csv"])
    assert len(df) == 3
    assert df["Name"].iloc[0] == "O'Reilly"


def test_export_with_unicode(temp_dir):
    """Test export handles Unicode characters."""
    data = pd.DataFrame(
        {
            "name": ["François", "北京", "Москва", "🚀 Rocket"],
            "value": [100, 200, 300, 400],
        }
    )

    builder = TableBuilder()
    config = builder.build(data, title="Unicode Test")

    result = export_table(config, temp_dir, formats=["csv", "excel"])

    # Verify CSV
    df_csv = pd.read_csv(result["csv"])
    assert df_csv["Name"].iloc[0] == "François"
    assert df_csv["Name"].iloc[1] == "北京"

    # Verify Excel
    df_excel = pd.read_excel(result["excel"])
    assert df_excel["Name"].iloc[2] == "Москва"


def test_export_with_nulls(temp_dir):
    """Test export handles null values correctly."""
    data = pd.DataFrame(
        {
            "col1": [1, None, 3, None],
            "col2": ["A", "B", None, "D"],
            "col3": [None, None, None, None],
        }
    )

    builder = TableBuilder()
    config = builder.build(data, title="Nulls Test")

    result = export_table(config, temp_dir, formats=["csv"])

    df = pd.read_csv(result["csv"])
    assert df["Col1"].isna().sum() == 2
    assert df["Col2"].isna().sum() == 1
    assert df["Col3"].isna().sum() == 4


def test_export_with_numeric_column_names(temp_dir):
    """Test export handles numeric column names."""
    # Convert numeric column names to strings first (pandas requirement)
    data = pd.DataFrame(
        {
            "1": [100, 200],
            "2": [300, 400],
            "3": [500, 600],
        }
    )

    builder = TableBuilder()
    config = builder.build(data, title="Numeric Columns")

    result = export_table(config, temp_dir, formats=["csv"])

    df = pd.read_csv(result["csv"])
    assert len(df.columns) == 3


def test_export_single_row(temp_dir):
    """Test export with single row of data."""
    data = pd.DataFrame({"col1": ["value1"], "col2": [123]})

    builder = TableBuilder()
    config = builder.build(data, title="Single Row")

    result = export_table(config, temp_dir, formats=["csv"])

    df = pd.read_csv(result["csv"])
    assert len(df) == 1


def test_export_single_column(temp_dir):
    """Test export with single column."""
    data = pd.DataFrame({"value": [1, 2, 3, 4, 5]})

    builder = TableBuilder()
    config = builder.build(data, title="Single Column")

    result = export_table(config, temp_dir, formats=["csv"])

    df = pd.read_csv(result["csv"])
    assert len(df.columns) == 1
    assert len(df) == 5


def test_export_very_long_strings(temp_dir):
    """Test export handles very long string values."""
    long_string = "A" * 10000  # 10k character string

    data = pd.DataFrame({"id": [1], "long_text": [long_string]})

    builder = TableBuilder()
    config = builder.build(data, title="Long Strings")

    result = export_table(config, temp_dir, formats=["csv", "excel"])

    # Verify both formats handle long strings
    df_csv = pd.read_csv(result["csv"])
    assert len(df_csv["Long Text"].iloc[0]) == 10000

    df_excel = pd.read_excel(result["excel"])
    assert len(df_excel["Long Text"].iloc[0]) == 10000
