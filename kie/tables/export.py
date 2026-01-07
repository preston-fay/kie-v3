"""
Table Export Utilities

Export tables to various formats (CSV, Excel, PDF).
"""

from pathlib import Path

import pandas as pd

from kie.tables.schema import TableConfig


class TableExporter:
    """Export tables to multiple formats."""

    def __init__(self):
        """Initialize table exporter."""
        pass

    def to_csv(
        self, config: TableConfig, output_path: str | Path, include_totals: bool = True
    ) -> Path:
        """
        Export table to CSV.

        Args:
            config: TableConfig
            output_path: Output file path
            include_totals: Include totals row if enabled

        Returns:
            Path to saved file
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Convert to DataFrame
        df = pd.DataFrame(config.data)

        # Filter to visible columns (handle empty DataFrame)
        visible_cols = [col.key for col in config.columns if not col.hidden]
        if len(df) > 0:
            df = df[visible_cols]
        else:
            # Empty DataFrame - create with correct columns
            df = pd.DataFrame(columns=visible_cols)

        # Rename columns to display names
        rename_map = {col.key: col.header for col in config.columns if not col.hidden}
        df = df.rename(columns=rename_map)

        # Add totals row if needed
        if include_totals and config.show_totals_row:
            totals_row = self._calculate_totals(config)
            df = pd.concat([df, pd.DataFrame([totals_row])], ignore_index=True)

        # Save
        df.to_csv(output_path, index=False)
        return output_path

    def to_excel(
        self,
        config: TableConfig,
        output_path: str | Path,
        sheet_name: str = "Data",
        include_totals: bool = True,
        style_headers: bool = True,
    ) -> Path:
        """
        Export table to Excel with formatting.

        Args:
            config: TableConfig
            output_path: Output file path
            sheet_name: Excel sheet name
            include_totals: Include totals row
            style_headers: Apply KDS styling to headers

        Returns:
            Path to saved file
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Convert to DataFrame
        df = pd.DataFrame(config.data)

        # Filter to visible columns (handle empty DataFrame)
        visible_cols = [col.key for col in config.columns if not col.hidden]
        if len(df) > 0:
            df = df[visible_cols]
        else:
            # Empty DataFrame - create with correct columns
            df = pd.DataFrame(columns=visible_cols)

        # Rename columns
        rename_map = {col.key: col.header for col in config.columns if not col.hidden}
        df = df.rename(columns=rename_map)

        # Add totals row
        if include_totals and config.show_totals_row:
            totals_row = self._calculate_totals(config)
            df = pd.concat([df, pd.DataFrame([totals_row])], ignore_index=True)

        # Write to Excel with formatting
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Style if requested
            if style_headers:
                self._style_excel(writer, sheet_name, config)

        return output_path

    def to_pdf(
        self, config: TableConfig, output_path: str | Path, include_totals: bool = True
    ) -> Path:
        """
        Export table to PDF with KDS styling.

        Args:
            config: TableConfig
            output_path: Output file path
            include_totals: Include totals row

        Returns:
            Path to saved file

        Raises:
            ImportError: If reportlab is not installed
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except ImportError:
            raise ImportError(
                "PDF export requires reportlab. Install with: pip install reportlab"
            )

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Convert to DataFrame
        df = pd.DataFrame(config.data)

        # Filter to visible columns (handle empty DataFrame)
        visible_cols = [col.key for col in config.columns if not col.hidden]
        if len(df) > 0:
            df = df[visible_cols]
        else:
            # Empty DataFrame - create with correct columns
            df = pd.DataFrame(columns=visible_cols)

        # Rename columns
        rename_map = {col.key: col.header for col in config.columns if not col.hidden}
        df = df.rename(columns=rename_map)

        # Add totals row
        if include_totals and config.show_totals_row:
            totals_row = self._calculate_totals(config)
            df = pd.concat([df, pd.DataFrame([totals_row])], ignore_index=True)

        # Determine page orientation based on column count
        page_size = landscape(letter) if len(df.columns) > 6 else letter

        # Create PDF document
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=page_size,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
            leftMargin=0.75 * inch,
            rightMargin=0.75 * inch,
        )

        # Build content
        story = []
        styles = getSampleStyleSheet()

        # Add title
        if config.title:
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#7823DC'),  # KDS purple
                spaceAfter=12,
                fontName='Helvetica-Bold'
            )
            title = Paragraph(config.title, title_style)
            story.append(title)
            story.append(Spacer(1, 0.2 * inch))

        # Prepare table data
        table_data = [df.columns.tolist()] + df.values.tolist()

        # Create table
        table = Table(table_data, repeatRows=1)

        # Apply KDS styling
        table_style = TableStyle([
            # Header row - KDS purple background, white text
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7823DC')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),

            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 1), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('LEFTPADDING', (0, 1), (-1, -1), 8),
            ('RIGHTPADDING', (0, 1), (-1, -1), 8),

            # Alternating row colors (light gray)
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),

            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D2D2D2')),

            # Outer border
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#4B4B4B')),
        ])

        # Style totals row if present
        if include_totals and config.show_totals_row:
            last_row_idx = len(table_data) - 1
            table_style.add('FONTNAME', (0, last_row_idx), (-1, last_row_idx), 'Helvetica-Bold')
            table_style.add('BACKGROUND', (0, last_row_idx), (-1, last_row_idx),
                          colors.HexColor('#E0E0E0'))

        table.setStyle(table_style)
        story.append(table)

        # Add footer with timestamp
        story.append(Spacer(1, 0.3 * inch))
        from datetime import datetime
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#787878'),
            alignment=2  # Right align
        )
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        footer = Paragraph(f"Generated by KIE v3 - {timestamp}", footer_style)
        story.append(footer)

        # Build PDF
        doc.build(story)
        return output_path

    def _calculate_totals(self, config: TableConfig) -> dict:
        """
        Calculate totals row.

        Args:
            config: TableConfig

        Returns:
            Dictionary with totals
        """
        totals = {}
        df = pd.DataFrame(config.data)

        for col in config.columns:
            if col.footer_aggregate:
                values = df[col.key].dropna()

                if col.footer_aggregate == "sum":
                    totals[col.header] = values.sum()
                elif col.footer_aggregate == "avg":
                    totals[col.header] = values.mean()
                elif col.footer_aggregate == "min":
                    totals[col.header] = values.min()
                elif col.footer_aggregate == "max":
                    totals[col.header] = values.max()
                elif col.footer_aggregate == "count":
                    totals[col.header] = len(values)
                else:
                    totals[col.header] = ""
            else:
                # First column gets "Total" label
                if col == config.columns[0]:
                    totals[col.header] = config.totals_label
                else:
                    totals[col.header] = ""

        return totals

    def _style_excel(self, writer, sheet_name: str, config: TableConfig):
        """
        Apply KDS styling to Excel worksheet.

        Args:
            writer: ExcelWriter
            sheet_name: Sheet name
            config: TableConfig
        """
        from openpyxl.styles import Alignment as ExcelAlignment
        from openpyxl.styles import Font, PatternFill

        worksheet = writer.sheets[sheet_name]

        # KDS purple for headers
        header_fill = PatternFill(start_color="7823DC", end_color="7823DC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", name="Inter")

        # Style header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = ExcelAlignment(horizontal="left", vertical="center")

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Style totals row if present
        if config.show_totals_row:
            last_row = worksheet.max_row
            totals_font = Font(bold=True, name="Inter")

            for cell in worksheet[last_row]:
                cell.font = totals_font


def export_table(
    config: TableConfig,
    output_dir: str | Path,
    formats: list = None,
    base_name: str | None = None,
) -> dict:
    """
    Export table to multiple formats.

    Args:
        config: TableConfig
        output_dir: Output directory
        formats: List of formats ('csv', 'excel', 'pdf')
        base_name: Base filename (uses title if None)

    Returns:
        Dictionary mapping format to file path
    """
    if formats is None:
        formats = ["csv", "excel"]
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Determine base filename
    if base_name is None:
        base_name = config.title.replace(" ", "_").lower() if config.title else "table"

    exporter = TableExporter()
    output_paths = {}

    for fmt in formats:
        if fmt == "csv":
            path = exporter.to_csv(config, output_dir / f"{base_name}.csv")
            output_paths["csv"] = path
        elif fmt == "excel":
            path = exporter.to_excel(config, output_dir / f"{base_name}.xlsx")
            output_paths["excel"] = path
        elif fmt == "pdf":
            try:
                path = exporter.to_pdf(config, output_dir / f"{base_name}.pdf")
                output_paths["pdf"] = path
            except ImportError:
                # reportlab not installed - skip PDF export silently
                pass

    return output_paths
